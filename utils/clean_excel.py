"""
sort_reviews_by_date.py
-----------------------
Processes every .xlsx file inside the `excel/` folder (in-place):

  1. Cleans 'Reviewer Metadata' → expands into:
         number_of_reviews | num_photos | is_local_guide
     (skipped if already done or column is absent)

  2. Sorts rows by the relative-date column (e.g. "3 months ago", "a year ago")
     so the latest reviews appear first.

Usage:
    python sort_reviews_by_date.py [options]

Options:
    --folder     Path to folder containing .xlsx files (default: excel)
    --col        Date column name or 0-based index (default: auto-detect)
    --sheet      Sheet name or 0-based index (default: 0 = first sheet)
    --no-clean   Skip the Reviewer Metadata cleaning step

Examples:
    python sort_reviews_by_date.py
    python sort_reviews_by_date.py --folder data/reviews
    python sort_reviews_by_date.py --col "Review Time Line"
    python sort_reviews_by_date.py --no-clean
"""

import os
import re
import sys
import argparse
from pathlib import Path
from copy import copy

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Reviewer Metadata cleaner
# ---------------------------------------------------------------------------

def parse_metadata(text):
    """Parse a 'Reviewer Metadata' cell into (num_reviews, num_photos, is_local_guide)."""
    if pd.isna(text):
        return 0, 0, False

    text = str(text).lower()
    is_local_guide = "local guide" in text

    reviews_match = re.search(r'(\d+)\s+review', text)
    num_reviews = int(reviews_match.group(1)) if reviews_match else 0

    photos_match = re.search(r'(\d+)\s+photo', text)
    num_photos = int(photos_match.group(1)) if photos_match else 0

    return num_reviews, num_photos, is_local_guide


def clean_reviewer_metadata(df):
    """Expand 'Reviewer Metadata' into three columns and drop the original."""
    DERIVED = {"number_of_reviews", "num_photos", "is_local_guide"}

    if DERIVED.issubset(df.columns):
        print("  [CLEAN] Already cleaned — skipping metadata expansion.")
        return df, False

    if "Reviewer Metadata" not in df.columns:
        print("  [CLEAN] 'Reviewer Metadata' column not found — skipping.")
        return df, False

    df[['number_of_reviews', 'num_photos', 'is_local_guide']] = (
        df['Reviewer Metadata'].apply(lambda x: pd.Series(parse_metadata(x)))
    )
    df.drop(columns=["Reviewer Metadata"], inplace=True)
    print("  [CLEAN] Expanded 'Reviewer Metadata' → number_of_reviews, num_photos, is_local_guide.")
    return df, True


# ---------------------------------------------------------------------------
# Relative-date parser
# ---------------------------------------------------------------------------

PATTERNS = [
    (r'^just\s+now$',               0),
    (r'^today$',                    0),
    (r'^yesterday$',                1),
    (r'^a\s+day\s+ago$',            1),
    (r'^(\d+)\s+day[s]?\s+ago$',   ('days', 1)),
    (r'^an?\s+hour[s]?\s+ago$',    0),
    (r'^(\d+)\s+hour[s]?\s+ago$',  0),
    (r'^a\s+week\s+ago$',           7),
    (r'^(\d+)\s+week[s]?\s+ago$',  ('days', 7)),
    (r'^a\s+month\s+ago$',          30),
    (r'^(\d+)\s+month[s]?\s+ago$', ('days', 30)),
    (r'^a\s+year\s+ago$',           365),
    (r'^(\d+)\s+year[s]?\s+ago$',  ('days', 365)),
]


def relative_to_days(text):
    """Convert a relative date string to approximate days. Lower = more recent."""
    if not isinstance(text, str):
        return float('inf')

    cleaned = text.strip().lower()
    cleaned = re.sub(r'^edited\s+', '', cleaned)

    for pattern, value in PATTERNS:
        m = re.match(pattern, cleaned)
        if m:
            if isinstance(value, tuple):
                _, multiplier = value
                return int(m.group(1)) * multiplier
            return value

    return float('inf')


# ---------------------------------------------------------------------------
# Auto-detect date column
# ---------------------------------------------------------------------------

DATE_KEYWORDS = re.compile(
    r'(review.?time|time.?line|date|posted|when|ago|review.?date)', re.I
)

RELATIVE_DATE_RE = re.compile(
    r'\b(just now|today|yesterday|\d+\s+(day|week|month|year|hour)s?\s+ago'
    r'|a\s+(day|week|month|year|hour)\s+ago)\b', re.I
)


def detect_date_column(df):
    """Return the column name most likely to contain relative review dates."""
    for col in df.columns:
        if DATE_KEYWORDS.search(str(col)):
            return col

    best_col, best_score = None, 0
    for col in df.columns:
        sample = df[col].dropna().astype(str).head(20)
        score = sum(1 for v in sample if RELATIVE_DATE_RE.search(v))
        if score > best_score:
            best_score, best_col = score, col

    return best_col if best_score > 0 else None


# ---------------------------------------------------------------------------
# Core: process a single file (in-place)
# ---------------------------------------------------------------------------

def process_file(file_path, col=None, sheet=0, clean=True):
    file_path = Path(file_path)

    df = pd.read_excel(file_path, sheet_name=sheet)

    if df.empty:
        print("  [SKIP] Sheet is empty.")
        return

    # Step 1: Clean
    if clean:
        df, _ = clean_reviewer_metadata(df)

    # Step 2: Resolve date column
    if col is not None:
        if isinstance(col, int):
            if col >= len(df.columns):
                print(f"  [SKIP] Column index {col} out of range ({len(df.columns)} columns).")
                return
            date_col = df.columns[col]
        else:
            if col not in df.columns:
                print(f"  [SKIP] Column '{col}' not found. Available: {list(df.columns)}")
                return
            date_col = col
    else:
        date_col = detect_date_column(df)
        if date_col is None:
            print(f"  [SKIP] Could not auto-detect a date column. Available: {list(df.columns)}")
            return
        print(f"  [SORT] Auto-detected date column: '{date_col}'")

    # Step 3: Sort
    df['_sort_key'] = df[date_col].apply(relative_to_days)
    df_sorted = df.sort_values('_sort_key', kind='stable').drop(columns=['_sort_key'])

    unrecognised = df[date_col][df['_sort_key'] == float('inf')].unique()
    if len(unrecognised):
        print(f"  [WARN] {len(unrecognised)} unrecognised date value(s) pushed to end: {list(unrecognised)}")

    # Step 4: Write back in-place, preserving header styles & column widths
    wb_orig = load_workbook(file_path)
    sheet_name = sheet if isinstance(sheet, str) else wb_orig.sheetnames[sheet]
    ws_orig = wb_orig[sheet_name]

    wb_new = Workbook()
    ws_new = wb_new.active
    ws_new.title = ws_orig.title

    orig_header_styles = {cell.value: cell for cell in ws_orig[1]}

    for col_idx, col_name in enumerate(df_sorted.columns, 1):
        new_cell = ws_new.cell(row=1, column=col_idx, value=col_name)
        orig_cell = orig_header_styles.get(col_name)
        if orig_cell:
            for attr in ('font', 'fill', 'alignment', 'border'):
                val = getattr(orig_cell, attr, None)
                if val:
                    setattr(new_cell, attr, copy(val))

    for new_row_idx, (_, row) in enumerate(df_sorted.iterrows(), 2):
        for col_idx, val in enumerate(row.values, 1):
            ws_new.cell(row=new_row_idx, column=col_idx, value=val)

    for col_idx in range(1, ws_orig.max_column + 1):
        letter = get_column_letter(col_idx)
        dim = ws_orig.column_dimensions.get(letter)
        ws_new.column_dimensions[letter].width = dim.width if dim else 20

    wb_new.save(file_path)
    print(f"  [DONE] Sorted {len(df_sorted)} rows — saved in-place.")


# ---------------------------------------------------------------------------
# Main: loop over all .xlsx files in the folder
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description='Clean + sort all .xlsx review files in a folder (in-place).'
    )
    parser.add_argument('--folder',   default='excel',
                        help='Folder containing .xlsx files (default: excel)')
    parser.add_argument('--col',      default=None,
                        help='Date column name or 0-based index (default: auto-detect)')
    parser.add_argument('--sheet',    default=0,
                        help='Sheet name or 0-based index (default: 0)')
    parser.add_argument('--no-clean', action='store_true',
                        help='Skip the Reviewer Metadata cleaning step')

    args = parser.parse_args()

    col = args.col
    if col is not None and col.isdigit():
        col = int(col)

    sheet = args.sheet
    if isinstance(sheet, str) and sheet.isdigit():
        sheet = int(sheet)

    folder = Path(args.folder)
    if not folder.is_dir():
        print(f"ERROR: Folder '{folder}' does not exist.")
        sys.exit(1)

    files = sorted(f for f in os.listdir(folder) if f.endswith('.xlsx'))
    if not files:
        print(f"No .xlsx files found in '{folder}'.")
        sys.exit(0)

    print(f"Found {len(files)} file(s) in '{folder}'\n")

    for filename in files:
        file_path = folder / filename
        print(f"→ {filename}")
        try:
            process_file(file_path, col=col, sheet=sheet, clean=not args.no_clean)
        except Exception as e:
            print(f"  [ERROR] {e}")
        print()


if __name__ == '__main__':
    main()