"""
sort_reviews_by_date.py
-----------------------
Processes every .xlsx file inside the `excel/` folder (in-place):

  1. Cleans 'Reviewer Metadata' → expands into:
         number_of_reviews | num_photos | is_local_guide

  2. Sorts rows by relative-date column (latest first)

  3. Extracts business name + location from filename:
         excel/<business_name>_<location>.xlsx
     e.g. excel/pita_shree_handicraft_udaipur.xlsx
          → business: "pita shree handicraft", location: "udaipur"

  4. Fetches Places API data and saves to:
         places_json/<stem>.json

Usage:
    python sort_reviews_by_date.py
    python sort_reviews_by_date.py --folder data/reviews
    python sort_reviews_by_date.py --no-clean
    python sort_reviews_by_date.py --skip-fetch
"""

import os
import re
import sys
import json
import argparse
from pathlib import Path
from copy import copy

import requests
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

load_dotenv(override=True)

GOOGLE_MAPS_API_KEY = os.getenv("GOOGLE_MAPS_API_KEY")
PLACES_SEARCH_URL   = "https://places.googleapis.com/v1/places:searchText"
PLACES_DETAILS_URL  = "https://places.googleapis.com/v1/places/{place_id}"


# ---------------------------------------------------------------------------
# Filename parser
# ---------------------------------------------------------------------------

def parse_filename(stem: str) -> tuple[str, str]:
    """
    Extract business name and location from filename stem.
    Last underscore-separated token = location, rest = business name.

    e.g. 'pita_shree_handicraft_udaipur' -> ('pita shree handicraft', 'udaipur')
    """
    parts = stem.split("_")
    if len(parts) == 1:
        return stem, ""
    location      = parts[-1]
    business_name = " ".join(parts[:-1])
    return business_name, location


# ---------------------------------------------------------------------------
# Reviewer Metadata cleaner
# ---------------------------------------------------------------------------

def parse_metadata(text):
    if pd.isna(text):
        return 0, 0, False
    text           = str(text).lower()
    is_local_guide = "local guide" in text
    reviews_match  = re.search(r'(\d+)\s+review', text)
    num_reviews    = int(reviews_match.group(1)) if reviews_match else 0
    photos_match   = re.search(r'(\d+)\s+photo', text)
    num_photos     = int(photos_match.group(1)) if photos_match else 0
    return num_reviews, num_photos, is_local_guide


def clean_reviewer_metadata(df):
    DERIVED = {"number_of_reviews", "num_photos", "is_local_guide"}
    if DERIVED.issubset(df.columns):
        print("  [CLEAN] Already cleaned — skipping.")
        return df, False
    if "Reviewer Metadata" not in df.columns:
        print("  [CLEAN] 'Reviewer Metadata' column not found — skipping.")
        return df, False
    df[['number_of_reviews', 'num_photos', 'is_local_guide']] = (
        df['Reviewer Metadata'].apply(lambda x: pd.Series(parse_metadata(x)))
    )
    df.drop(columns=["Reviewer Metadata"], inplace=True)
    print("  [CLEAN] Expanded Reviewer Metadata → number_of_reviews, num_photos, is_local_guide.")
    return df, True


# ---------------------------------------------------------------------------
# Relative-date parser
# ---------------------------------------------------------------------------

PATTERNS = [
    (r'^just\s+now$',               0),
    (r'^today$',                    0),
    (r'^yesterday$',                1),
    (r'^a\s+day\s+ago$',            1),
    (r'^(\d+)\s+day[s]?\s+ago$',   ('days', 1)),
    (r'^an?\s+hour[s]?\s+ago$',    0),
    (r'^(\d+)\s+hour[s]?\s+ago$',  0),
    (r'^a\s+week\s+ago$',           7),
    (r'^(\d+)\s+week[s]?\s+ago$',  ('days', 7)),
    (r'^a\s+month\s+ago$',          30),
    (r'^(\d+)\s+month[s]?\s+ago$', ('days', 30)),
    (r'^a\s+year\s+ago$',           365),
    (r'^(\d+)\s+year[s]?\s+ago$',  ('days', 365)),
]


def relative_to_days(text):
    if not isinstance(text, str):
        return float('inf')
    cleaned = re.sub(r'^edited\s+', '', text.strip().lower())
    for pattern, value in PATTERNS:
        m = re.match(pattern, cleaned)
        if m:
            if isinstance(value, tuple):
                _, multiplier = value
                return int(m.group(1)) * multiplier
            return value
    return float('inf')


# ---------------------------------------------------------------------------
# Auto-detect date column
# ---------------------------------------------------------------------------

DATE_KEYWORDS    = re.compile(r'(review.?time|time.?line|date|posted|when|ago|review.?date)', re.I)
RELATIVE_DATE_RE = re.compile(
    r'\b(just now|today|yesterday|\d+\s+(day|week|month|year|hour)s?\s+ago'
    r'|a\s+(day|week|month|year|hour)\s+ago)\b', re.I
)


def detect_date_column(df):
    for col in df.columns:
        if DATE_KEYWORDS.search(str(col)):
            return col
    best_col, best_score = None, 0
    for col in df.columns:
        sample = df[col].dropna().astype(str).head(20)
        score  = sum(1 for v in sample if RELATIVE_DATE_RE.search(v))
        if score > best_score:
            best_score, best_col = score, col
    return best_col if best_score > 0 else None


# ---------------------------------------------------------------------------
# Places API fetch + save
# ---------------------------------------------------------------------------

def fetch_and_save(business_name: str, location: str, output_stem: str, category: str = None):
    if not GOOGLE_MAPS_API_KEY:
        print("  [SKIP FETCH] GOOGLE_MAPS_API_KEY not set in .env")
        return

    print(f"  [FETCH] Searching: '{business_name}' near '{location}'")

    query   = f"{business_name} {location}".strip()
    headers = {
        "Content-Type": "application/json",
        "X-Goog-Api-Key": GOOGLE_MAPS_API_KEY,
        "X-Goog-FieldMask": "places.id,places.displayName,places.formattedAddress",
    }
    response = requests.post(PLACES_SEARCH_URL, headers=headers, json={"textQuery": query})
    response.raise_for_status()
    places = response.json().get("places", [])
    if not places:
        print("  [FETCH] No matching business found.")
        return

    place    = places[0]
    place_id = place.get("id")
    name     = place.get("displayName", {}).get("text", "N/A")
    address  = place.get("formattedAddress", "N/A")
    print(f"  [FETCH] Found: {name} — {address}")

    url = PLACES_DETAILS_URL.format(place_id=place_id)
    headers = {
        "X-Goog-Api-Key": GOOGLE_MAPS_API_KEY,
        "X-Goog-FieldMask": (
            "displayName,formattedAddress,nationalPhoneNumber,"
            "internationalPhoneNumber,websiteUri,rating,userRatingCount,"
            "regularOpeningHours,businessStatus"
        ),
    }
    details  = requests.get(url, headers=headers).json()
    if "error" in details:
        print(f"  [FETCH ERROR] {details['error'].get('message')}")
        return

    phone    = details.get("nationalPhoneNumber") or details.get("internationalPhoneNumber", "N/A")
    website  = details.get("websiteUri", "N/A")
    rating   = details.get("rating", "N/A")
    count    = details.get("userRatingCount", "N/A")
    status   = details.get("businessStatus", "N/A")
    hours    = details.get("regularOpeningHours", {})
    open_now = hours.get("openNow")
    open_str = "Open Now" if open_now is True else "Closed Now" if open_now is False else "N/A"

    info = {
        "name":               name,
        "address":            address,
        "category":           category if category else "Unknown",
        "phone_number":       phone,
        "website":            website,
        "average_rating":     rating,
        "total_rating_count": count,
        "business_status":    status,
        "open_now":           open_str,
    }

    out_dir = Path("places_json")
    out_dir.mkdir(exist_ok=True)
    out_path = out_dir / f"{output_stem}.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(info, f, indent=2, ensure_ascii=False)
    print(f"  [SAVED] {out_path}")


# ---------------------------------------------------------------------------
# Core: process a single file
# ---------------------------------------------------------------------------

def process_file(file_path: Path, col=None, sheet=0, clean=True, skip_fetch=False, category=None):
    print(f"\n-> {file_path.name}")

    business_name, location = parse_filename(file_path.stem)
    print(f"  [INFO] Business: '{business_name}' | Location: '{location}'")

    df = pd.read_excel(file_path, sheet_name=sheet)
    if df.empty:
        print("  [SKIP] Sheet is empty.")
        return

    # Clean
    if clean:
        df, _ = clean_reviewer_metadata(df)

    # Resolve date column
    if col is not None:
        date_col = df.columns[col] if isinstance(col, int) else col
        if date_col not in df.columns:
            print(f"  [SKIP] Column '{date_col}' not found.")
            return
    else:
        date_col = detect_date_column(df)
        if date_col:
            print(f"  [SORT] Auto-detected date column: '{date_col}'")
        else:
            print("  [WARN] Could not detect date column — skipping sort.")

    # Sort
    if date_col:
        df['_sort_key'] = df[date_col].apply(relative_to_days)
        df = df.sort_values('_sort_key', kind='stable').drop(columns=['_sort_key'])
        print(f"  [SORT] Sorted {len(df)} rows by recency.")

    # Write back in-place preserving header styles + column widths
    wb_orig    = load_workbook(file_path)
    sheet_name = sheet if isinstance(sheet, str) else wb_orig.sheetnames[sheet]
    ws_orig    = wb_orig[sheet_name]
    wb_new     = Workbook()
    ws_new     = wb_new.active
    ws_new.title = ws_orig.title

    orig_styles = {cell.value: cell for cell in ws_orig[1]}
    for col_idx, col_name in enumerate(df.columns, 1):
        new_cell  = ws_new.cell(row=1, column=col_idx, value=col_name)
        orig_cell = orig_styles.get(col_name)
        if orig_cell:
            for attr in ('font', 'fill', 'alignment', 'border'):
                val = getattr(orig_cell, attr, None)
                if val:
                    setattr(new_cell, attr, copy(val))

    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        for col_idx, val in enumerate(row.values, 1):
            ws_new.cell(row=row_idx, column=col_idx, value=val)

    for col_idx in range(1, ws_orig.max_column + 1):
        letter = get_column_letter(col_idx)
        dim    = ws_orig.column_dimensions.get(letter)
        ws_new.column_dimensions[letter].width = dim.width if dim else 20

    wb_new.save(file_path)
    print(f"  [DONE] Saved cleaned + sorted file in-place.")

    # Fetch Places API
    if not skip_fetch:
        fetch_and_save(business_name, location, file_path.stem, category=category)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description='Clean + sort a .xlsx review file and fetch Places API data.',
        epilog="""
Examples:
  python sort_reviews_by_date.py excel/pita_shree_handicraft_udaipur.xlsx
  python sort_reviews_by_date.py excel/pita_shree_handicraft_udaipur.xlsx --skip-fetch
  python sort_reviews_by_date.py excel/pita_shree_handicraft_udaipur.xlsx --no-clean
        """
    )
    parser.add_argument('filepath',
                        help='Path to the .xlsx file (e.g. excel/pita_shree_handicraft_udaipur.xlsx)')
    parser.add_argument('--col',        default=None,
                        help='Date column name or 0-based index (default: auto-detect)')
    parser.add_argument('--sheet',      default=0,
                        help='Sheet name or 0-based index (default: 0)')
    parser.add_argument('--no-clean',   action='store_true',
                        help='Skip the Reviewer Metadata cleaning step')
    parser.add_argument('--skip-fetch', action='store_true',
                        help='Skip the Places API fetch step')
    parser.add_argument('--category',   default=None,
                        help='Business category (e.g. Restaurant, Retail, Pharmacy)')

    args      = parser.parse_args()
    col       = int(args.col) if args.col and args.col.isdigit() else args.col
    sheet     = int(args.sheet) if isinstance(args.sheet, str) and args.sheet.isdigit() else args.sheet
    file_path = Path(args.filepath)

    if not file_path.exists():
        print(f"ERROR: File '{file_path}' does not exist.")
        sys.exit(1)

    try:
        process_file(
            file_path,
            col=col,
            sheet=sheet,
            clean=not args.no_clean,
            skip_fetch=args.skip_fetch,
            category=args.category,
        )
    except Exception as e:
        print(f"  [ERROR] {e}")

    print("\nAll done.")


if __name__ == '__main__':
    main()