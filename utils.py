"""
utils.py
--------
Shared utilities for the BRS scoring pipeline.

Contains:
  - RECENCY_MAP         : relative date string → approximate days
  - recency_multiplier  : days → recency weight
  - parse_filename      : stem → (business_name, location)
  - write_back_excel    : write dataframe back to xlsx preserving styles
"""

import json
from copy import copy
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Recency map — relative date string → approximate days
# ---------------------------------------------------------------------------

RECENCY_MAP = {
    "Just now": 0, "a minute ago": 0, "2 minutes ago": 0,
    "3 minutes ago": 0, "4 minutes ago": 0, "5 minutes ago": 0,
    "10 minutes ago": 0, "15 minutes ago": 0, "20 minutes ago": 0,
    "30 minutes ago": 0, "45 minutes ago": 0,
    "an hour ago": 0, "2 hours ago": 0, "3 hours ago": 0,
    "4 hours ago": 0, "5 hours ago": 0, "6 hours ago": 0,
    "7 hours ago": 0, "8 hours ago": 0, "9 hours ago": 0,
    "10 hours ago": 0, "11 hours ago": 0, "12 hours ago": 0,
    "13 hours ago": 0, "14 hours ago": 0, "15 hours ago": 0,
    "16 hours ago": 0, "17 hours ago": 0, "18 hours ago": 0,
    "19 hours ago": 0, "20 hours ago": 0, "21 hours ago": 0,
    "22 hours ago": 0, "23 hours ago": 0,
    "yesterday": 1,
    "2 days ago": 2,  "3 days ago": 3,  "4 days ago": 4,
    "5 days ago": 5,  "6 days ago": 6,
    "a week ago": 7,  "2 weeks ago": 14,
    "3 weeks ago": 21, "4 weeks ago": 28,
    "a month ago": 30,  "2 months ago": 60,  "3 months ago": 90,
    "4 months ago": 120, "5 months ago": 150, "6 months ago": 180,
    "7 months ago": 210, "8 months ago": 240, "9 months ago": 270,
    "10 months ago": 300, "11 months ago": 330,
    "a year ago": 365,  "2 years ago": 730,  "3 years ago": 1095,
    "4 years ago": 1460, "5 years ago": 1825, "6 years ago": 2190,
    "7 years ago": 2555, "8 years ago": 2920, "9 years ago": 3285,
    "10 years ago": 3650,
}


def relative_to_days(text: str) -> float:
    """
    Convert a relative date string to approximate days.
    Strips 'Edited' prefix if present.
    Returns float('inf') if unrecognised.
    """
    if not isinstance(text, str):
        return float("inf")
    cleaned = text.strip().lower()
    if cleaned.startswith("edited "):
        cleaned = cleaned[len("edited "):]
    return RECENCY_MAP.get(cleaned, float("inf"))


# ---------------------------------------------------------------------------
# Recency multiplier — days → weight (0.0 to 1.0)
# ---------------------------------------------------------------------------

def recency_multiplier(days: float) -> float:
    """
    Map approximate days to a recency weight.

    Thresholds:
      ≤ 90 days  → 1.00  (< 3 months)
      ≤ 365 days → 0.85  (3–12 months)
      ≤ 730 days → 0.65  (1–2 years)
      ≤ 1095     → 0.45  (2–3 years)
      > 1095     → 0.25  (3+ years)
    """
    if days <= 90:   return 1.00
    if days <= 365:  return 0.85
    if days <= 730:  return 0.65
    if days <= 1095: return 0.45
    return 0.25


# ---------------------------------------------------------------------------
# Filename parser — stem → (business_name, location)
# ---------------------------------------------------------------------------

def parse_filename(stem: str) -> tuple[str, str]:
    """
    Extract business name and location from filename stem.
    Convention: <business_name>_<location>
    Last underscore-separated token = location, rest = business name.

    e.g. 'pita_shree_handicraft_udaipur' → ('pita shree handicraft', 'udaipur')
    """
    parts = stem.split("_")
    if len(parts) == 1:
        return stem, ""
    location      = parts[-1]
    business_name = " ".join(parts[:-1])
    return business_name, location


# ---------------------------------------------------------------------------
# Load places JSON
# ---------------------------------------------------------------------------

def load_places_json(stem: str, folder: str = "places_json") -> dict:
    """
    Load Places API data from places_json/<stem>.json.
    Returns empty dict if file not found.
    """
    path = Path(folder) / f"{stem}.json"
    if not path.exists():
        return {}
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


# ---------------------------------------------------------------------------
# Write dataframe back to xlsx preserving header styles + column widths
# ---------------------------------------------------------------------------

def write_back_excel(file_path: Path, df: pd.DataFrame, sheet: int | str = 0):
    """
    Write df back to file_path in-place, preserving:
      - Header cell styles (font, fill, alignment, border)
      - Column widths
    """
    wb_orig    = load_workbook(file_path)
    sheet_name = sheet if isinstance(sheet, str) else wb_orig.sheetnames[sheet]
    ws_orig    = wb_orig[sheet_name]
    wb_new     = Workbook()
    ws_new     = wb_new.active
    ws_new.title = ws_orig.title

    orig_styles = {cell.value: cell for cell in ws_orig[1]}
    for col_idx, col_name in enumerate(df.columns, 1):
        new_cell  = ws_new.cell(row=1, column=col_idx, value=col_name)
        orig_cell = orig_styles.get(col_name)
        if orig_cell:
            for attr in ("font", "fill", "alignment", "border"):
                val = getattr(orig_cell, attr, None)
                if val:
                    setattr(new_cell, attr, copy(val))

    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        for col_idx, val in enumerate(row.values, 1):
            ws_new.cell(row=row_idx, column=col_idx, value=val)

    for col_idx in range(1, ws_orig.max_column + 1):
        letter = get_column_letter(col_idx)
        dim    = ws_orig.column_dimensions.get(letter)
        ws_new.column_dimensions[letter].width = dim.width if dim else 20

    wb_new.save(file_path)